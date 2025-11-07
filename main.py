# main.py — финальная версия: тихое игнорирование неавторизованных, предупреждение на /start, комментарии на русском
import json
import os
import sys
import asyncio
import hashlib
import time
import re
import csv
import traceback
from datetime import datetime
from cryptography.fernet import Fernet
from dotenv import load_dotenv
from telethon import TelegramClient
from telethon.errors import SessionPasswordNeededError, RPCError
from telethon.tl.types import ChannelParticipantsAdmins
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes
)

# попытка импортировать openpyxl для создания xlsx
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

# -----------------------
# === Логи и конфигурация ===
# -----------------------
LOG_DIR = "logs"
if not os.path.exists(LOG_DIR):
    os.makedirs(LOG_DIR)

LOG_SESSIONS = os.path.join(LOG_DIR, "logging.log")
LOG_ERRORS = os.path.join(LOG_DIR, "error.log")
LOG_WRONG_ACCESS = os.path.join(LOG_DIR, "wrong_access.log")


def now_iso_local():
    """Текущее локальное время в ISO 8601 с часовым поясом."""
    return datetime.now().astimezone().isoformat()


def log_session(msg: str):
    """Запись событий по созданию/удалению сессий."""
    try:
        with open(LOG_SESSIONS, "a", encoding="utf-8") as f:
            f.write(f"{now_iso_local()} | {msg}\n")
    except Exception:
        pass


def log_error(exc_text: str):
    """Запись полного traceback ошибок в error.log."""
    try:
        with open(LOG_ERRORS, "a", encoding="utf-8") as f:
            f.write(f"{now_iso_local()} | {exc_text}\n\n")
    except Exception:
        pass


def log_wrong_access(user_id: int, msg: str):
    """Запись попыток доступа неавторизованных пользователей."""
    try:
        with open(LOG_WRONG_ACCESS, "a", encoding="utf-8") as f:
            f.write(f"{now_iso_local()} | user_id={user_id} | {msg}\n")
    except Exception:
        pass


# -----------------------
# === Переменные окружения / .env ===
# -----------------------
ENV_PATH = ".env"
if not os.path.exists(ENV_PATH):
    with open(ENV_PATH, "w", encoding="utf-8") as f:
        f.write("API_ID=11986433\n")
        f.write("API_HASH=17af9a3324d57bfb85f0a7ac8b98a60d\n")

load_dotenv(ENV_PATH)
try:
    API_ID = int(os.getenv("API_ID"))
except Exception:
    API_ID = None
API_HASH = os.getenv("API_HASH")

# -----------------------
# === Аргументы запуска (токен бота) ===
# -----------------------
if len(sys.argv) < 2:
    print("Использование: python main.py <BOT_TOKEN>")
    sys.exit(1)
BOT_TOKEN = sys.argv[1]

# -----------------------
# === Пользователи / шифрованный файл ===
# -----------------------
USERS_FILE = "users.enc"
KEY_FILE = "secret.key"


def generate_key():
    """Генерация ключа Fernet и запись в KEY_FILE."""
    key = Fernet.generate_key()
    with open(KEY_FILE, "wb") as f:
        f.write(key)


def load_key():
    """Загрузка ключа Fernet; если нет — генерируем новый."""
    if not os.path.exists(KEY_FILE):
        generate_key()
    with open(KEY_FILE, "rb") as f:
        return f.read()


fernet = Fernet(load_key())


def load_users():
    """Загрузка users_data из зашифрованного файла users.enc."""
    if not os.path.exists(USERS_FILE):
        save_users({"admins": [], "operators": []})
    try:
        with open(USERS_FILE, "rb") as f:
            data = f.read()
            if not data:
                return {"admins": [], "operators": []}
            return json.loads(fernet.decrypt(data).decode())
    except Exception:
        log_error("Failed to load users.enc:\n" + traceback.format_exc())
        return {"admins": [], "operators": []}


def save_users(data):
    """Сохранение users_data в зашифрованный файл users.enc."""
    try:
        with open(USERS_FILE, "wb") as f:
            f.write(fernet.encrypt(json.dumps(data).encode()))
    except Exception:
        log_error("Failed to save users.enc:\n" + traceback.format_exc())


users_data = load_users()

# -----------------------
# === Временные структуры в памяти ===
# -----------------------
pending_action = {}  # словарь текущих flow для пользователей
# active_sessions: map session_path -> {"created": ts, "expiry": ts, "owner": user_id}
active_sessions = {}

# глобальная ссылка на приложение для фоновых задач и удаления сообщений
GLOBAL_APP = None

# TTL сессии (в секундах)
SESSION_TTL_SECONDS = 15 * 60  # 15 минут

# -----------------------
# === Метки кнопок / UI ===
# -----------------------
LABEL_ADD_ADMIN = "➕ Добавить администратора"
LABEL_ADD_OPERATOR = "➕ Добавить оператора"
LABEL_REMOVE_OPERATOR = "➖ Удалить оператора"
LABEL_LIST_OPERATORS = "📋 Список операторов"
LABEL_SCAN = "Сканировать"
LABEL_CANCEL = "Отмена"

# -----------------------
# === Функции для клавиатур ===
# -----------------------
def main_menu_keyboard(role: str):
    """Клавиатура главного меню в зависимости от роли."""
    if role == "admin":
        return ReplyKeyboardMarkup(
            [
                [LABEL_ADD_ADMIN, LABEL_ADD_OPERATOR, LABEL_REMOVE_OPERATOR],
                [LABEL_LIST_OPERATORS, LABEL_SCAN]
            ],
            resize_keyboard=True
        )
    elif role == "operator":
        return ReplyKeyboardMarkup([[LABEL_SCAN]], resize_keyboard=True)
    else:
        return ReplyKeyboardMarkup([["/start"]], resize_keyboard=True)


def cancel_keyboard():
    """Клавиатура с одной кнопкой Отмена."""
    return ReplyKeyboardMarkup([[LABEL_CANCEL]], resize_keyboard=True)


def chats_keyboard(chats):
    """Генерация клавиатуры из списка названий чатов (макс 25) плюс Отмена."""
    buttons = [[chat] for chat in chats[:25]]
    buttons.append([LABEL_CANCEL])
    return ReplyKeyboardMarkup(buttons, resize_keyboard=True)


# -----------------------
# === Утилиты / нормализация ===
# -----------------------
def normalize(text: str) -> str:
    """Нормализация текста для сравнения меток: trim, убираем ведущий '/', приводим к нижнему регистру."""
    if not text:
        return ""
    t = text.strip()
    if t.startswith("/"):
        t = t[1:]
    parts = t.split()
    return " ".join(parts).lower()


def is_same_label(text: str, label: str) -> bool:
    """Проверка эквивалентности метки с нормализацией."""
    return normalize(text) == normalize(label)


def get_user_role(user_id: int):
    """Возвращает роль пользователя по user_id: 'admin' | 'operator' | None."""
    users = users_data
    if user_id in users.get("admins", []):
        return "admin"
    if user_id in users.get("operators", []):
        return "operator"
    return None


async def send_main_menu(update: Update, user_id: int):
    """Отправка главного меню пользователю (без проверок)."""
    role = get_user_role(user_id)
    try:
        await update.message.reply_text("Возврат в главное меню:", reply_markup=main_menu_keyboard(role))
    except Exception:
        log_error("Failed to send main menu:\n" + traceback.format_exc())


# -----------------------
# === Telethon: сессии и авторизация ===
# -----------------------
SESSIONS_DIR = "sessions"
if not os.path.exists(SESSIONS_DIR):
    os.makedirs(SESSIONS_DIR)


def normalize_phone(raw: str) -> str:
    """
    Универсальная нормализация телефона:
    - 8xxxxxxxxxx -> +7xxxxxxxxxx
    - 10 цифр -> +7xxxxxxxxxx
    - сохраняем ведущий '+' если есть
    - fallback: '+' + digits
    """
    if not raw:
        return raw
    raw = raw.strip()
    plus_prefixed = raw.startswith("+")
    digits = re.sub(r"\D", "", raw)
    if plus_prefixed:
        return "+" + digits
    if len(digits) == 11 and digits.startswith("8"):
        return "+7" + digits[1:]
    if len(digits) == 11 and digits.startswith("7"):
        return "+" + digits
    if len(digits) == 10:
        return "+7" + digits
    if len(digits) > 0:
        return "+" + digits
    return raw


def _session_filename_for_phone(phone: str):
    """Генерация имени и пути сессионного файла по sha256(phone)."""
    h = hashlib.sha256(phone.encode()).hexdigest()
    session_name = f"session_{h}"
    session_filepath = os.path.join(SESSIONS_DIR, session_name + ".session")
    return session_name, session_filepath


async def telethon_send_code(phone: str):
    """Отправка кода через Telethon (send_code_request). Возвращает (client, None) или (None, err)."""
    if API_ID is None or not API_HASH:
        raise RuntimeError("API_ID/API_HASH не заданы в .env")

    session_name, session_filepath = _session_filename_for_phone(phone)
    if os.path.exists(session_filepath):
        return None, "exists"

    client = TelegramClient(os.path.join(SESSIONS_DIR, session_name), API_ID, API_HASH)
    try:
        await client.connect()
        await client.send_code_request(phone)
        log_session(f"Sent code request for phone {phone} (session={session_filepath})")
        return client, None
    except Exception:
        log_error("telethon_send_code error:\n" + traceback.format_exc())
        try:
            await client.disconnect()
        except Exception:
            pass
        return None, "send_code_error"


async def list_user_chats_and_store(client: TelegramClient, update: Update, user_id: int):
    """
    Получаем диалоги (только супергруппы/каналы, пригодные для сканирования),
    сохраняем их в pending_action[user_id]['dialogs'] и показываем список.
    После успешного получения чатов — удаляем все auth-сообщения.
    """
    try:
        dialogs = await client.get_dialogs(limit=200)
    except Exception:
        log_error("Error in get_dialogs:\n" + traceback.format_exc())
        try:
            await update.message.reply_text("Ошибка при получении чатов.", reply_markup=main_menu_keyboard(get_user_role(user_id)))
        except Exception:
            log_error("Failed to inform user about get_dialogs error:\n" + traceback.format_exc())
        return

    dialogs_filtered = []
    titles = []
    for d in dialogs:
        title = None
        if hasattr(d, "title") and d.title:
            title = d.title
        elif hasattr(d, "name") and d.name:
            title = d.name
        else:
            try:
                title = getattr(d.entity, "title", None) or getattr(d.entity, "name", None)
            except Exception:
                title = None

        is_channel = getattr(d, "is_channel", False)
        is_group = getattr(d, "is_group", False)
        if (is_group or is_channel) and title:
            dialogs_filtered.append({"title": title, "id": getattr(d, "id", None), "dialog": d})
            titles.append(title)

    if not titles:
        try:
            await update.message.reply_text("⚠️ Не найдено доступных чатов.", reply_markup=main_menu_keyboard(get_user_role(update.effective_user.id)))
        except Exception:
            log_error("Failed to reply 'no chats':\n" + traceback.format_exc())
        try:
            await client.disconnect()
        except Exception:
            pass
        return

    existing = pending_action.get(update.effective_user.id, {})
    auth_msgs = existing.get("auth_messages", [])
    pending_action[update.effective_user.id] = {
        "action": "choose_chat",
        "phone": existing.get("phone"),
        "client": client,
        "dialogs": dialogs_filtered,
        "auth_messages": auth_msgs,
        "start_time": existing.get("start_time", time.time())
    }

    # после формирования списка чатов — пытаемся удалить auth сообщения (и ботовые, и пользовательские)
    try:
        await purge_auth_messages_for_user(update.effective_user.id)
    except Exception:
        log_error("Failed to purge auth messages after listing chats:\n" + traceback.format_exc())

    try:
        await update.message.reply_text("Выберите чат:", reply_markup=chats_keyboard(titles))
    except Exception:
        log_error("Failed to send chat list message:\n" + traceback.format_exc())


# -----------------------
# === Экспорт участников в XLSX/CSV ===
# -----------------------
def safe_filename(s: str) -> str:
    """Создание безопасного имени файла из строки (удаляем нежелательные символы)."""
    return "".join(c if c.isalnum() or c in " _-()" else "_" for c in s)[:120]


async def export_members_to_xlsx_and_send(client: TelegramClient, dialog, requester_chat_id: int, bot_app):
    """
    Собираем до 100 участников, формируем XLSX с автошириной колонок,
    отправляем пользователю и удаляем файл.
    Колонки: TelegramID, Status, Username (@...), FullName, Phone, JoinedDate (YYYY-MM-DD)
    """
    members = []
    try:
        entity = dialog.entity
    except Exception:
        entity = dialog

    try:
        admins = await client.get_participants(entity, filter=ChannelParticipantsAdmins())
        admin_ids = {u.id for u in admins}
    except Exception:
        admin_ids = set()

    cnt = 0
    try:
        async for user in client.iter_participants(entity, limit=100):
            uid = getattr(user, "id", "")
            username = getattr(user, "username", "") or ""
            if username and not username.startswith("@"):
                username = "@" + username
            fname = getattr(user, "first_name", "") or ""
            lname = getattr(user, "last_name", "") or ""
            full_name = (fname + " " + lname).strip()
            phone = getattr(user, "phone", "") or ""
            phone_display = normalize_phone(phone) if phone else ""
            joined = ""
            try:
                part = getattr(user, "participant", None)
                if part is not None:
                    joined_attr = getattr(part, "date", None)
                    if joined_attr:
                        try:
                            joined = joined_attr.strftime("%Y-%m-%d")
                        except Exception:
                            joined = str(joined_attr)
            except Exception:
                joined = ""

            status = "Admin" if uid in admin_ids else "User"

            members.append({
                "TelegramID": uid,
                "Status": status,
                "Username": username,
                "FullName": full_name,
                "Phone": phone_display,
                "JoinedDate": joined
            })
            cnt += 1
            if cnt >= 100:
                break
    except Exception:
        log_error("Error while iterating participants:\n" + traceback.format_exc())
        try:
            await bot_app.bot.send_message(chat_id=requester_chat_id, text=f"❌ Error while fetching participants.")
        except Exception:
            log_error("Failed to notify user about participant fetch error:\n" + traceback.format_exc())
        return

    if not members:
        try:
            await bot_app.bot.send_message(chat_id=requester_chat_id, text="(No members found in selected group)")
        except Exception:
            log_error("Failed to notify no members found:\n" + traceback.format_exc())
        return

    dialog_id = getattr(dialog, "id", int(time.time()))
    safe_title = safe_filename(getattr(dialog, "title", None) or getattr(dialog, "name", "") or str(dialog_id))
    ts = int(time.time())
    xlsx_filename = f"chat_members_{dialog_id}_{ts}.xlsx"
    csv_fallback = f"chat_members_{dialog_id}_{ts}.csv"

    if OPENPYXL_AVAILABLE:
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Members"

            headers = ["TelegramID", "Status", "Username", "FullName", "Phone", "JoinedDate"]
            ws.append(headers)

            max_lens = [len(h) for h in headers]
            for row in members:
                row_values = [row["TelegramID"], row["Status"], row["Username"], row["FullName"], row["Phone"], row["JoinedDate"]]
                ws.append(row_values)
                for i, v in enumerate(row_values):
                    s = "" if v is None else str(v)
                    l = len(s)
                    if l > max_lens[i]:
                        max_lens[i] = l

            for i, width in enumerate(max_lens, start=1):
                col_letter = get_column_letter(i)
                calc_width = min(max(8, int(width * 1.1) + 2), 80)
                ws.column_dimensions[col_letter].width = calc_width

            wb.save(xlsx_filename)
            await bot_app.bot.send_document(chat_id=requester_chat_id, document=open(xlsx_filename, "rb"))
        except Exception:
            log_error("XLSX generation/sending failed:\n" + traceback.format_exc())
            try:
                await bot_app.bot.send_message(chat_id=requester_chat_id, text="⚠️ XLSX failed, falling back to CSV.")
            except Exception:
                log_error("Failed to notify about XLSX fallback:\n" + traceback.format_exc())
            try:
                with open(csv_fallback, "w", newline="", encoding="utf-8-sig") as csvfile:
                    fieldnames = ["TelegramID", "Status", "Username", "FullName", "Phone", "JoinedDate"]
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter=";", quoting=csv.QUOTE_ALL)
                    writer.writeheader()
                    for r in members:
                        writer.writerow(r)
                await bot_app.bot.send_document(chat_id=requester_chat_id, document=open(csv_fallback, "rb"))
            except Exception:
                log_error("Failed to send fallback CSV:\n" + traceback.format_exc())
            finally:
                try:
                    if os.path.exists(csv_fallback):
                        os.remove(csv_fallback)
                except Exception:
                    pass
        finally:
            try:
                if os.path.exists(xlsx_filename):
                    os.remove(xlsx_filename)
            except Exception:
                pass
    else:
        try:
            await bot_app.bot.send_message(chat_id=requester_chat_id, text="⚠️ 'openpyxl' not installed — sending CSV instead. To enable XLSX install: pip install openpyxl")
        except Exception:
            pass
        try:
            with open(csv_fallback, "w", newline="", encoding="utf-8-sig") as csvfile:
                fieldnames = ["TelegramID", "Status", "Username", "FullName", "Phone", "JoinedDate"]
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter=";", quoting=csv.QUOTE_ALL)
                writer.writeheader()
                for r in members:
                    writer.writerow(r)
            await bot_app.bot.send_document(chat_id=requester_chat_id, document=open(csv_fallback, "rb"))
        except Exception:
            log_error("Failed to create/send csv fallback:\n" + traceback.format_exc())
        finally:
            try:
                if os.path.exists(csv_fallback):
                    os.remove(csv_fallback)
            except Exception:
                pass


# -----------------------
# === Удаление сообщений авторизации ===
# -----------------------
async def purge_auth_messages_for_user(user_id: int):
    """
    Удаление сообщений, записанных в pending_action[user_id]['auth_messages'].
    Удаляем как пользовательские сообщения, так и подсказки/ответы бота (DEL:2).
    Ошибки игнорируем, но логируем.
    """
    global GLOBAL_APP
    if GLOBAL_APP is None:
        return
    pa = pending_action.get(user_id)
    if not pa:
        return
    msgs = pa.get("auth_messages", [])
    for m in msgs:
        chat_id = m.get("chat_id")
        message_id = m.get("message_id")
        try:
            await GLOBAL_APP.bot.delete_message(chat_id=chat_id, message_id=message_id)
        except Exception:
            log_error(f"Failed to delete message {message_id} in chat {chat_id} for user {user_id}:\n" + traceback.format_exc())
    pa["auth_messages"] = []


def record_auth_message(user_id: int, chat_id: int, message_id: int, from_bot: bool):
    """Заносим сообщение (user или bot) в список для последующего удаления."""
    pa = pending_action.get(user_id)
    entry = {"chat_id": chat_id, "message_id": message_id, "from_bot": bool(from_bot)}
    if pa is None:
        pending_action[user_id] = {"auth_messages": [entry], "start_time": time.time()}
    else:
        lst = pa.get("auth_messages", [])
        lst.append(entry)
        pa["auth_messages"] = lst
        if "start_time" not in pa:
            pa["start_time"] = time.time()


# -----------------------
# === Обработчики команд ===
# -----------------------
# Текст предупреждения (вариант A) — показываем каждому авторизованному пользователю на /start
WARNING_TEXT = "⚠️ Внимание! Использование аккаунта Telegram для сканирования чатов может привести к его блокировке. Вы используете это на свой риск."


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обработчик /start:
    - если пользователь авторизован (admin/operator) — показываем предупреждение и главное меню;
    - если пользователь не авторизован — молча игнорируем (но логируем попытку).
    """
    user_id = update.effective_user.id
    role = get_user_role(user_id)
    if not role:
        # Молчаливое игнорирование — только лог
        log_wrong_access(user_id, "Attempted /start without access")
        return

    # показываем предупреждение (каждый /start) и затем главное меню
    try:
        # сначала предупреждение
        await update.message.reply_text(WARNING_TEXT)
    except Exception:
        log_error("Failed to send warning:\n" + traceback.format_exc())

    try:
        await update.message.reply_text("Добро пожаловать в главное меню:", reply_markup=main_menu_keyboard(role))
    except Exception:
        log_error("Failed to send main menu on start:\n" + traceback.format_exc())


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Команда Отмена:
    - если пользователь в pending_action — отключаем client, очищаем pending, удаляем auth-сообщения;
    - затем возвращаем пользователя в главное меню.
    """
    user_id = update.effective_user.id
    role = get_user_role(user_id)
    if not role:
        # молча игнорируем, логируем
        log_wrong_access(user_id, "Attempted /cancel without access")
        return

    if user_id in pending_action:
        client = pending_action[user_id].get("client")
        try:
            if client:
                await client.disconnect()
        except Exception:
            log_error("Error disconnecting client at cancel:\n" + traceback.format_exc())
        try:
            await purge_auth_messages_for_user(user_id)
        except Exception:
            log_error("Failed to purge auth messages at cancel:\n" + traceback.format_exc())
        pending_action.pop(user_id, None)
    await send_main_menu(update, user_id)


# -----------------------
# === Обработчик текстовых сообщений ===
# -----------------------
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Главный обработчик текстовых сообщений:
    - если пользователь не авторизован — молча игнорируем (только логируем);
    - если пользователь авторизован — обрабатываем поток (scan, add/remove admin/operator и т.д.).
    """
    user_id = update.effective_user.id
    raw_text = (update.message.text or "")
    text = raw_text.strip()
    n = normalize(text)
    role = get_user_role(user_id)

    # если пользователь не авторизован — молча игнорируем и логируем
    if role is None:
        log_wrong_access(user_id, f"Unauthorized message: {text[:200]}")
        return

    # если пользователь в pending flow — записываем входящее сообщение для последующего удаления
    if user_id in pending_action:
        try:
            record_auth_message(user_id, update.effective_chat.id, update.message.message_id, from_bot=False)
        except Exception:
            log_error("Failed to record auth message (user msg):\n" + traceback.format_exc())

    # Обработка команды Отмена через кнопку
    if n == normalize(LABEL_CANCEL):
        if user_id in pending_action:
            client = pending_action[user_id].get("client")
            try:
                if client:
                    await client.disconnect()
            except Exception:
                log_error("Error disconnecting client at cancel (handle_message):\n" + traceback.format_exc())
            try:
                await purge_auth_messages_for_user(user_id)
            except Exception:
                log_error("Failed to purge auth messages at cancel (handle_message):\n" + traceback.format_exc())
            pending_action.pop(user_id, None)
        await send_main_menu(update, user_id)
        return

    # Если пользователь в pending flow — обрабатываем соответствующие шаги
    if user_id in pending_action:
        action = pending_action[user_id]
        act = action.get("action")

        # flow логина/сканирования
        if act == "login":
            step = action.get("step")
            if step == "phone":
                phone_raw = text
                if not phone_raw or (not phone_raw[0].isdigit() and phone_raw[0] != "+" and phone_raw[0] != "8"):
                    try:
                        msg = await update.message.reply_text("❌ Некорректный номер. Введите в формате +79161234567 или нажмите Отмена.", reply_markup=cancel_keyboard())
                        record_auth_message(user_id, msg.chat.id, msg.message_id, from_bot=True)
                    except Exception:
                        log_error("Failed sending invalid phone message:\n" + traceback.format_exc())
                    return
                phone_norm = normalize_phone(phone_raw)
                pa = pending_action.get(user_id, {})
                pa["phone"] = phone_norm
                pa["start_time"] = pa.get("start_time", time.time())
                pending_action[user_id] = pa

                session_name, session_path = _session_filename_for_phone(phone_norm)
                # если сессия уже есть — используем её и показываем список чатов
                if os.path.exists(session_path):
                    try:
                        client = TelegramClient(os.path.join(SESSIONS_DIR, session_name), API_ID, API_HASH)
                        await client.connect()
                        now = time.time()
                        active_sessions[session_path] = {"created": now, "expiry": now + SESSION_TTL_SECONDS, "owner": user_id}
                        log_session(f"Re-used session activated for {phone_norm} (path={session_path}) by user {user_id}")
                        await list_user_chats_and_store(client, update, user_id)
                    except Exception:
                        log_error("Failed to use existing session:\n" + traceback.format_exc())
                        try:
                            await client.disconnect()
                        except Exception:
                            pass
                    return

                # иначе отправляем код
                try:
                    msg = await update.message.reply_text("⏳ Отправляем код подтверждения...", reply_markup=cancel_keyboard())
                    record_auth_message(user_id, msg.chat.id, msg.message_id, from_bot=True)
                except Exception:
                    log_error("Failed to send 'sending code' message:\n" + traceback.format_exc())
                client, err = await telethon_send_code(phone_norm)
                if client is None:
                    if err == "exists":
                        try:
                            await update.message.reply_text("⚠️ Сессия уже существует.", reply_markup=main_menu_keyboard(role))
                        except Exception:
                            log_error("Failed to reply 'session exists':\n" + traceback.format_exc())
                    else:
                        try:
                            await update.message.reply_text(f"❌ Ошибка при отправке кода: {err}", reply_markup=main_menu_keyboard(role))
                        except Exception:
                            log_error("Failed to reply 'send code error':\n" + traceback.format_exc())
                    pending_action.pop(user_id, None)
                    return
                try:
                    msg2 = await update.message.reply_text("📩 Код подтверждения отправлен. Введите его:", reply_markup=cancel_keyboard())
                    record_auth_message(user_id, msg2.chat.id, msg2.message_id, from_bot=True)
                except Exception:
                    log_error("Failed to send 'enter code' prompt:\n" + traceback.format_exc())
                pending_action[user_id] = {"action": "login", "step": "code", "client": client, "phone": phone_norm, "start_time": time.time(), "auth_messages": pending_action[user_id].get("auth_messages", [])}
                return

            if step == "code":
                code = text
                client = action.get("client")
                phone = action.get("phone")
                try:
                    await client.sign_in(phone, code)
                    session_name, session_path = _session_filename_for_phone(phone)
                    now = time.time()
                    active_sessions[session_path] = {"created": now, "expiry": now + SESSION_TTL_SECONDS, "owner": user_id}
                    log_session(f"Session created for {phone} (path={session_path}) by user {user_id}")
                    await list_user_chats_and_store(client, update, user_id)
                    return
                except SessionPasswordNeededError:
                    pending_action[user_id] = {"action": "login", "step": "password", "client": client, "phone": phone, "start_time": action.get("start_time", time.time()), "auth_messages": action.get("auth_messages", [])}
                    try:
                        msg = await update.message.reply_text("🔒 Введите пароль 2FA:", reply_markup=cancel_keyboard())
                        record_auth_message(user_id, msg.chat.id, msg.message_id, from_bot=True)
                    except Exception:
                        log_error("Failed to send 2FA prompt:\n" + traceback.format_exc())
                    return
                except Exception:
                    log_error("Error in sign_in code step:\n" + traceback.format_exc())
                    try:
                        await client.disconnect()
                    except Exception:
                        pass
                    pending_action.pop(user_id, None)
                    try:
                        await update.message.reply_text(f"❌ Ошибка авторизации.", reply_markup=main_menu_keyboard(role))
                    except Exception:
                        log_error("Failed to reply auth error message:\n" + traceback.format_exc())
                    return

            if step == "password":
                password = text
                client = action.get("client")
                phone = action.get("phone")
                try:
                    await client.sign_in(password=password)
                    session_name, session_path = _session_filename_for_phone(phone)
                    now = time.time()
                    active_sessions[session_path] = {"created": now, "expiry": now + SESSION_TTL_SECONDS, "owner": user_id}
                    log_session(f"Session created (2FA) for {phone} (path={session_path}) by user {user_id}")
                    await list_user_chats_and_store(client, update, user_id)
                    return
                except Exception:
                    log_error("Error in sign_in password step:\n" + traceback.format_exc())
                    try:
                        await client.disconnect()
                    except Exception:
                        pass
                    pending_action.pop(user_id, None)
                    try:
                        await update.message.reply_text(f"❌ Ошибка 2FA.", reply_markup=main_menu_keyboard(role))
                    except Exception:
                        log_error("Failed to reply 2FA error:\n" + traceback.format_exc())
                    return

        # выбор чата из списка
        if act == "choose_chat":
            dialogs = action.get("dialogs", [])
            matched = None
            for d in dialogs:
                if d["title"] == text:
                    matched = d
                    break
            if not matched:
                try:
                    await update.message.reply_text("❌ Чат не найден. Нажмите Отмена и попробуйте снова.", reply_markup=chats_keyboard([d["title"] for d in dialogs]))
                except Exception:
                    log_error("Failed to send 'chat not found' message:\n" + traceback.format_exc())
                return
            dialog_obj = matched.get("dialog")
            client = action.get("client")
            try:
                await export_members_to_xlsx_and_send(client, dialog_obj, update.effective_chat.id, context.application)
            except Exception:
                log_error("Error exporting members:\n" + traceback.format_exc())
                try:
                    await context.application.bot.send_message(chat_id=update.effective_chat.id, text="❌ Ошибка при экспорте участников.")
                except Exception:
                    log_error("Failed to notify user about export error:\n" + traceback.format_exc())
            try:
                await client.disconnect()
            except Exception:
                log_error("Failed to disconnect client after export:\n" + traceback.format_exc())
            pending_action.pop(user_id, None)
            await send_main_menu(update, user_id)
            return

    # -----------------------
    # === Главное меню: обработка кнопок ===
    # -----------------------
    if normalize(text) == normalize(LABEL_SCAN):
        if role not in ["admin", "operator"]:
            log_wrong_access(user_id, f"Tried to use Scan without rights")
            return
        pending_action[user_id] = {"action": "login", "step": "phone", "start_time": time.time(), "auth_messages": []}
        try:
            await update.message.reply_text("📱 Введите номер телефона для входа:", reply_markup=cancel_keyboard())
        except Exception:
            log_error("Failed to send 'enter phone' prompt:\n" + traceback.format_exc())
        return

    if normalize(text) == normalize(LABEL_ADD_ADMIN):
        if role != "admin":
            log_wrong_access(user_id, f"Tried to add admin without rights")
            return
        pending_action[user_id] = {"action": "add_admin"}
        await update.message.reply_text("Введите Telegram ID нового администратора:", reply_markup=cancel_keyboard())
        return

    if normalize(text) == normalize(LABEL_ADD_OPERATOR):
        if role != "admin":
            log_wrong_access(user_id, f"Tried to add operator without rights")
            return
        pending_action[user_id] = {"action": "add_operator"}
        await update.message.reply_text("Введите Telegram ID нового оператора:", reply_markup=cancel_keyboard())
        return

    if normalize(text) == normalize(LABEL_REMOVE_OPERATOR):
        if role != "admin":
            log_wrong_access(user_id, f"Tried to remove operator without rights")
            return
        ops = users_data.get("operators", [])
        if not ops:
            await update.message.reply_text("Операторов нет.", reply_markup=main_menu_keyboard(role))
            return
        buttons = [[str(x)] for x in ops]
        buttons.append([LABEL_CANCEL])
        pending_action[user_id] = {"action": "remove_operator"}
        await update.message.reply_text("Выберите оператора для удаления (нажмите ID):", reply_markup=ReplyKeyboardMarkup(buttons, resize_keyboard=True))
        return

    if normalize(text) == normalize(LABEL_LIST_OPERATORS):
        if role != "admin":
            log_wrong_access(user_id, f"Tried to list operators without rights")
            return
        ops = users_data.get("operators", [])
        if not ops:
            await update.message.reply_text("Операторов нет.", reply_markup=main_menu_keyboard(role))
            return
        await update.message.reply_text("Список операторов:\n" + "\n".join(str(x) for x in ops), reply_markup=main_menu_keyboard(role))
        return

    # неизвестная команда
    try:
        await update.message.reply_text("Неизвестная команда. Используйте кнопки меню.")
    except Exception:
        log_error("Failed to send 'unknown command' message:\n" + traceback.format_exc())


# -----------------------
# === Фоновая очистка сессий и pending ===
# -----------------------
async def session_and_pending_cleaner():
    """
    Фоновая задача:
    - удаляет сессии, у которых истек TTL от момента создания (политика A),
    - удаляет pending_action старше TTL,
    - пытается дисконнектить клиентов и удалять auth-сообщения.
    Запускается каждые 30 секунд.
    """
    global active_sessions, pending_action
    while True:
        now = time.time()
        # Удаляем устаревшие сессии по времени создания
        to_remove_sessions = []
        for session_path, meta in list(active_sessions.items()):
            created = meta.get("created", 0)
            if now >= (created + SESSION_TTL_SECONDS):
                session_base = os.path.splitext(session_path)[0]
                try:
                    for fname in os.listdir(SESSIONS_DIR):
                        if fname.startswith(os.path.basename(session_base)):
                            fpath = os.path.join(SESSIONS_DIR, fname)
                            try:
                                os.remove(fpath)
                                log_session(f"Removed expired session file: {fpath}")
                            except Exception:
                                log_error("Failed to remove session file in cleaner:\n" + traceback.format_exc())
                except Exception:
                    log_error("Error scanning sessions dir in cleaner:\n" + traceback.format_exc())
                to_remove_sessions.append(session_path)
        for p in to_remove_sessions:
            active_sessions.pop(p, None)

        # Очищаем зависшие pending flows старше TTL
        stale_users = []
        for uid, pa in list(pending_action.items()):
            start = pa.get("start_time")
            if start and (now - start) > SESSION_TTL_SECONDS:
                client = pa.get("client")
                try:
                    if client:
                        await client.disconnect()
                except Exception:
                    log_error("Error disconnecting client during pending cleanup:\n" + traceback.format_exc())
                try:
                    await purge_auth_messages_for_user(uid)
                except Exception:
                    log_error("Failed to purge auth messages during pending cleanup:\n" + traceback.format_exc())
                stale_users.append(uid)
        for uid in stale_users:
            pending_action.pop(uid, None)

        await asyncio.sleep(30)


# -----------------------
# === Supervisor и запуск приложения ===
# -----------------------
def main():
    global GLOBAL_APP
    if not os.path.exists(SESSIONS_DIR):
        os.makedirs(SESSIONS_DIR)

    # создаём приложение один раз
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    GLOBAL_APP = app

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("cancel", cancel))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    # старт фоновой задачи после инициализации app
    async def _start_background_tasks(app):
        try:
            asyncio.create_task(session_and_pending_cleaner())
        except Exception:
            log_error("Failed to start background tasks:\n" + traceback.format_exc())

    app.post_init = _start_background_tasks

    # супервайзорный цикл: перезапуск при исключении с паузой 5 секунд
    while True:
        try:
            print("Бот запускается. Если он упадет, supervisor перезапустит его через 5 сек.")
            app.run_polling()
        except Exception:
            log_error("App.run_polling crashed:\n" + traceback.format_exc())
            time.sleep(5)
            continue
        else:
            break


if __name__ == "__main__":
    main()
