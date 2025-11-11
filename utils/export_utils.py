# export_utils.py — функции экспорта участников
# Изменения: вынесены функции экспорта участников из main.py

import os
import re
import time
import csv
from datetime import datetime


# попытка импортировать openpyxl для создания xlsx
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

from telethon.tl.types import ChannelParticipantsAdmins
from utils.logging_utils import log_error


def safe_filename(s: str) -> str:
    return "".join(c if c.isalnum() or c in " _-()" else "_" for c in s)[:120]


async def export_members_to_xlsx_and_send(client, dialog, requester_chat_id: int, bot_app):
    members = []
    try:
        entity = dialog.entity
    except Exception:
        entity = dialog

    try:
        admins = await client.get_participants(entity, filter=ChannelParticipantsAdmins())
        admin_ids = {u.id for u in admins}
    except Exception:
        admin_ids = set()

    cnt = 0
    try:
        async for user in client.iter_participants(entity, limit=10):
            uid = getattr(user, "id", "")
            username = getattr(user, "username", "") or ""
            if username and not username.startswith("@"):
                username = "@" + username
            fname = getattr(user, "first_name", "") or ""
            lname = getattr(user, "last_name", "") or ""
            full_name = (fname + " " + lname).strip()
            phone = getattr(user, "phone", "") or ""
            phone_display = normalize_phone(phone) if phone else ""  # Эта функция должна быть определена или импортирована
            joined = ""
            try:
                part = getattr(user, "participant", None)
                if part is not None:
                    joined_attr = getattr(part, "date", None)
                    if joined_attr:
                        try:
                            joined = joined_attr.strftime("%Y-%m-%d")
                        except Exception:
                            joined = str(joined_attr)
            except Exception:
                joined = ""

            status = "Admin" if uid in admin_ids else "User"

            members.append({
                "TelegramID": uid,
                "Status": status,
                "Username": username,
                "FullName": full_name,
                "Phone": phone_display,
                "JoinedDate": joined
            })
            cnt += 1
            if cnt >= 100:
                break
    except Exception:
        log_error("Ошибка при переборе участников:\n" + str(locals()))
        try:
            await bot_app.bot.send_message(chat_id=requester_chat_id, text=f"❌ Ошибка при получении участников.")
        except Exception:
            log_error("Не удалось уведомить пользователя об ошибке получения участников:\n" + str(locals()))
        return

    if not members:
        try:
            await bot_app.bot.send_message(chat_id=requester_chat_id, text="(В выбранной группе нет участников)")
        except Exception:
            log_error("Не удалось уведомить об отсутствии участников:\n" + str(locals()))
        return

    dialog_id = getattr(dialog, "id", int(time.time()))
    safe_title = safe_filename(getattr(dialog, "title", None) or getattr(dialog, "name", "") or str(dialog_id))
    ts = int(time.time())
    xlsx_filename = f"chat_members_{dialog_id}_{ts}.xlsx"
    csv_fallback = f"chat_members_{dialog_id}_{ts}.csv"

    if OPENPYXL_AVAILABLE:
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Members"

            headers = ["TelegramID", "Status", "Username", "FullName", "Phone", "JoinedDate"]
            ws.append(headers)

            max_lens = [len(h) for h in headers]
            for row in members:
                row_values = [row["TelegramID"], row["Status"], row["Username"], row["FullName"], row["Phone"], row["JoinedDate"]]
                ws.append(row_values)
                for i, v in enumerate(row_values):
                    s = "" if v is None else str(v)
                    l = len(s)
                    if l > max_lens[i]:
                        max_lens[i] = l

            for i, width in enumerate(max_lens, start=1):
                col_letter = get_column_letter(i)
                calc_width = min(max(8, int(width * 1.1) + 2), 80)
                ws.column_dimensions[col_letter].width = calc_width

            wb.save(xlsx_filename)
            await bot_app.bot.send_document(chat_id=requester_chat_id, document=open(xlsx_filename, "rb"))
        except Exception:
            log_error("Создание/отправка XLSX не удалось:\n" + str(locals()))
            try:
                await bot_app.bot.send_message(chat_id=requester_chat_id, text="⚠️ XLSX не удалось, резервный вариант — CSV.")
            except Exception:
                log_error("Не удалось уведомить о резервном варианте XLSX:\n" + str(locals()))
            try:
                with open(csv_fallback, "w", newline="", encoding="utf-8-sig") as csvfile:
                    fieldnames = ["TelegramID", "Status", "Username", "FullName", "Phone", "JoinedDate"]
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter=";", quoting=csv.QUOTE_ALL)
                    writer.writeheader()
                    for r in members:
                        writer.writerow(r)
                await bot_app.bot.send_document(chat_id=requester_chat_id, document=open(csv_fallback, "rb"))
            except Exception:
                log_error("Не удалось отправить резервный CSV:\n" + str(locals()))
            finally:
                try:
                    if os.path.exists(csv_fallback):
                        os.remove(csv_fallback)
                except Exception:
                    pass
        finally:
            try:
                if os.path.exists(xlsx_filename):
                    os.remove(xlsx_filename)
            except Exception:
                pass
    else:
        try:
            await bot_app.bot.send_message(chat_id=requester_chat_id, text="⚠️ 'openpyxl' не установлен — отправка CSV вместо этого. Для включения XLSX установите: pip install openpyxl")
        except Exception:
            pass
        try:
            with open(csv_fallback, "w", newline="", encoding="utf-8-sig") as csvfile:
                fieldnames = ["TelegramID", "Status", "Username", "FullName", "Phone", "JoinedDate"]
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter=";", quoting=csv.QUOTE_ALL)
                writer.writeheader()
                for r in members:
                    writer.writerow(r)
            await bot_app.bot.send_document(chat_id=requester_chat_id, document=open(csv_fallback, "rb"))
        except Exception:
            log_error("Не удалось создать/отправить резервный csv:\n" + str(locals()))
        finally:
            try:
                if os.path.exists(csv_fallback):
                    os.remove(csv_fallback)
            except Exception:
                pass


def normalize_phone(raw: str) -> str:
    if not raw:
        return raw
    raw = raw.strip()
    plus_prefixed = raw.startswith("+")
    digits = re.sub(r"\D", "", raw)
    if plus_prefixed:
        return "+" + digits
    if len(digits) == 11 and digits.startswith("8"):
        return "+7" + digits[1:]
    if len(digits) == 11 and digits.startswith("7"):
        return "+" + digits
    if len(digits) == 10:
        return "+7" + digits
    if digits > 0:
        return "+" + digits
    return raw
